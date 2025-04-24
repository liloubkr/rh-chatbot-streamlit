import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl import load_workbook
import numpy as np
from datetime import datetime

# Configuration
FILE_PATH = "Data Reporting KPI RH Q32024.xlsx"
RECRUITERS = ["Inès", "Mariéme", "Pauline", "Samya"]
MONTHS = ["Juillet", "Août", "Septembre"]
KPI_OPTIONS = {
    "Nb de candidats contactés": "Candidats contactés",
    "Nb d'entretiens": "Entretiens (salariés + sous-traitants)",
    "Nb de recrutements": "Recrutements aboutis",
    "Nb d'entretiens salariés": "Entretiens avec salariés",
    "Nb d'entretiens sous-traitants": "Entretiens avec sous-traitants"
}

# Fonction pour extraire les données (identique à précédemment)
def extract_recruiter_data(sheet_name):
    df = pd.read_excel(FILE_PATH, sheet_name=sheet_name, header=None)
    
    data_start = None
    for i, row in df.iterrows():
        if row.str.contains("RECRUTEMENT").any():
            data_start = i
            break
    
    if data_start is None:
        raise ValueError(f"Données RECRUTEMENT non trouvées pour {sheet_name}")
    
    data = {}
    for i in range(data_start, data_start + 10):
        row = df.iloc[i]
        kpi_name = row[1]
        
        if pd.isna(kpi_name):
            continue
        
        kpi_name = kpi_name.strip()
        values = []
        
        for month_idx, month in enumerate(MONTHS):
            if month_idx == 0:
                val = row[3]
            elif month_idx == 1:
                val = row[8]
            elif month_idx == 2:
                val = row[9]
            
            if isinstance(val, str) and val.startswith('='):
                wb = load_workbook(FILE_PATH, data_only=True)
                ws = wb[sheet_name]
                cell_ref = ws.cell(row=i+1, column=9+month_idx).value
                val = cell_ref if cell_ref is not None else 0
            elif pd.isna(val):
                val = 0
            
            values.append(float(val) if val not in [None, np.nan] else 0)
        
        total = row[10]
        if isinstance(total, str) and total.startswith('='):
            wb = load_workbook(FILE_PATH, data_only=True)
            ws = wb[sheet_name]
            cell_ref = ws.cell(row=i+1, column=11).value
            total = cell_ref if cell_ref is not None else sum(values)
        elif pd.isna(total):
            total = sum(values)
        
        values.append(float(total) if total not in [None, np.nan] else sum(values))
        data[kpi_name] = values
    
    return pd.DataFrame(data, index=MONTHS + ["Total"])

# Chargement des données
@st.cache_data
def load_all_data():
    all_data = {}
    for recruiter in RECRUITERS:
        try:
            all_data[recruiter] = extract_recruiter_data(recruiter)
        except Exception as e:
            st.error(f"Erreur lors du chargement des données pour {recruiter}: {str(e)}")
            continue
    return all_data

# Calcul des KPI globaux
def calculate_kpis(all_data):
    results = {
        'total_contacted': 0,
        'total_interviews': 0,
        'total_salarie_interviews': 0,
        'total_sous_traitant_interviews': 0,
        'total_recruitments': 0,
        'recruiters_data': {}
    }
    
    best_contact = {"name": "", "value": 0}
    best_recruitment = {"name": "", "value": 0}
    
    for name, data in all_data.items():
        contacted = data.get("Nb de candidats contactés", [0]*4)
        salarie = data.get("Nb d'entretiens candidats Salariés", [0]*4)
        sous_traitant = data.get("Nb d'entretiens candidats Sous-Traitants", [0]*4)
        rec_salarie = data.get("Nb de candidats recrutés Salariés", [0]*4)
        rec_sous_traitant = data.get("Nb de candidats intégrés Sous Traitants", [0]*4)
        
        total_contacted = contacted[-1] if len(contacted) > 3 else sum(contacted[:3])
        total_salarie = salarie[-1] if len(salarie) > 3 else sum(salarie[:3])
        total_sous_traitant = sous_traitant[-1] if len(sous_traitant) > 3 else sum(sous_traitant[:3])
        total_rec = (rec_salarie[-1] if len(rec_salarie) > 3 else sum(rec_salarie[:3])) + \
                   (rec_sous_traitant[-1] if len(rec_sous_traitant) > 3 else sum(rec_sous_traitant[:3]))
        
        results['total_contacted'] += total_contacted
        results['total_salarie_interviews'] += total_salarie
        results['total_sous_traitant_interviews'] += total_sous_traitant
        results['total_interviews'] += total_salarie + total_sous_traitant
        results['total_recruitments'] += total_rec
        
        results['recruiters_data'][name] = {
            'Candidats contactés': total_contacted,
            'Entretiens Salariés': total_salarie,
            'Entretiens Sous-Traitants': total_sous_traitant,
            'Recrutements': total_rec
        }
        
        if total_contacted > best_contact["value"]:
            best_contact = {"name": name, "value": total_contacted}
        
        if total_rec > best_recruitment["value"]:
            best_recruitment = {"name": name, "value": total_rec}
    
    results['best_contact'] = best_contact
    results['best_recruitment'] = best_recruitment
    
    return results

# Fonctions de visualisation
def plot_kpi_trend(all_data, kpi_name, selected_recruiters=None):
    if selected_recruiters is None:
        selected_recruiters = RECRUITERS
    
    plt.figure(figsize=(10, 5))
    
    for recruiter in selected_recruiters:
        if recruiter in all_data and kpi_name in all_data[recruiter].columns:
            values = all_data[recruiter].loc[MONTHS, kpi_name]
            plt.plot(MONTHS, values, marker='o', label=recruiter)
    
    plt.title(f'Évolution du {kpi_name} par mois')
    plt.xlabel('Mois')
    plt.ylabel('Quantité')
    plt.legend()
    plt.grid(True, linestyle='--', alpha=0.7)
    st.pyplot(plt)

def plot_recruiter_comparison(all_data, kpi_name):
    recruiters = []
    values = []
    
    for recruiter, data in all_data.items():
        if kpi_name in data.columns:
            recruiters.append(recruiter)
            values.append(data.loc["Total", kpi_name])
    
    if not recruiters:
        st.warning(f"Aucune donnée disponible pour {kpi_name}")
        return
    
    plt.figure(figsize=(10, 5))
    plt.bar(recruiters, values)
    plt.title(f'Comparaison des recruteurs - {kpi_name} (total trimestre)')
    plt.ylabel('Quantité')
    plt.xticks(rotation=45)
    st.pyplot(plt)

# Interface du chatbot
def main():
    st.set_page_config(page_title="Chatbot RH - Reporting Q3 2024", page_icon="📊")
    
    # Initialisation de l'état de la conversation
    if "messages" not in st.session_state:
        st.session_state.messages = []
        st.session_state.messages.append({
            "role": "assistant", 
            "content": "Bonjour! Je suis votre assistant pour consulter les données RH du 3ème trimestre 2024. Que souhaitez-vous savoir?"
        })
    
    # Chargement des données
    all_data = load_all_data()
    kpis = calculate_kpis(all_data)
    
    # Affichage de l'historique de la conversation
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
    
    # Gestion de l'entrée utilisateur
    if prompt := st.chat_input("Posez votre question..."):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)
        
        # Préparation de la réponse
        response = ""
        lower_prompt = prompt.lower()
        
        # Réponses prédéfinies
        if any(word in lower_prompt for word in ["bonjour", "salut", "hello", "coucou"]):
            response = "Bonjour! Comment puis-je vous aider avec les données RH du Q3 2024?"
        
        elif "total" in lower_prompt:
            if "candidats contactés" in lower_prompt:
                response = f"Nombre total de candidats contactés: {kpis['total_contacted']}"
            elif "entretiens" in lower_prompt and "salariés" in lower_prompt:
                response = f"Nombre total d'entretiens avec salariés: {kpis['total_salarie_interviews']}"
            elif "entretiens" in lower_prompt and "sous-traitants" in lower_prompt:
                response = f"Nombre total d'entretiens avec sous-traitants: {kpis['total_sous_traitant_interviews']}"
            elif "entretiens" in lower_prompt:
                response = f"Nombre total d'entretiens: {kpis['total_interviews']}"
            elif "recrutements" in lower_prompt:
                response = f"Nombre total de recrutements aboutis: {kpis['total_recruitments']}"
            else:
                response = "Voici les totaux globaux:\n"
                response += f"- Candidats contactés: {kpis['total_contacted']}\n"
                response += f"- Entretiens totaux: {kpis['total_interviews']}\n"
                response += f"  - dont salariés: {kpis['total_salarie_interviews']}\n"
                response += f"  - dont sous-traitants: {kpis['total_sous_traitant_interviews']}\n"
                response += f"- Recrutements aboutis: {kpis['total_recruitments']}"
        
        elif "meilleur" in lower_prompt or "top" in lower_prompt:
            if "contacté" in lower_prompt or "contactés" in lower_prompt:
                response = f"Le recruteur ayant contacté le plus de candidats: {kpis['best_contact']['name']} ({kpis['best_contact']['value']} candidats)"
            elif "recruté" in lower_prompt or "recrutements" in lower_prompt:
                response = f"Le recruteur ayant effectué le plus de recrutements: {kpis['best_recruitment']['name']} ({kpis['best_recruitment']['value']} recrutements)"
            else:
                response = "Voici les meilleurs recruteurs:\n"
                response += f"- Contact: {kpis['best_contact']['name']} ({kpis['best_contact']['value']} candidats)\n"
                response += f"- Recrutements: {kpis['best_recruitment']['name']} ({kpis['best_recruitment']['value']} recrutements)"
        
        elif "graphique" in lower_prompt or "visualisation" in lower_prompt or "courbe" in lower_prompt:
            selected_kpi = st.selectbox("Choisissez l'indicateur à visualiser:", list(KPI_OPTIONS.keys()))
            
            if selected_kpi:
                response = f"Voici l'évolution du {selected_kpi} par mois:"
                
                # Déterminer quels recruteurs inclure
                include_all = st.checkbox("Tous les recruteurs", value=True)
                selected_recruiters = RECRUITERS if include_all else st.multiselect(
                    "Sélectionnez les recruteurs:",
                    RECRUITERS,
                    default=RECRUITERS
                )
                
                st.markdown(response)
                plot_kpi_trend(all_data, selected_kpi, selected_recruiters)
                return
        
        elif "comparaison" in lower_prompt or "comparer" in lower_prompt:
            selected_kpi = st.selectbox("Choisissez l'indicateur à comparer:", list(KPI_OPTIONS.keys()))
            
            if selected_kpi:
                response = f"Comparaison des recruteurs pour {selected_kpi}:"
                st.markdown(response)
                plot_recruiter_comparison(all_data, selected_kpi)
                return
        
        elif "aide" in lower_prompt or "soutien" in lower_prompt:
            response = "Je peux vous aider avec:\n"
            response += "- Les totaux globaux (candidats, entretiens, recrutements)\n"
            response += "- Les performances par recruteur\n"
            response += "- Les meilleurs recruteurs par catégorie\n"
            response += "- Des graphiques d'évolution mensuelle\n"
            response += "- Des comparaisons entre recruteurs\n\n"
            response += "Exemples de questions:\n"
            response += "- 'Quel est le total des candidats contactés?'\n"
            response += "- 'Qui a effectué le plus de recrutements?'\n"
            response += "- 'Montrez-moi un graphique des entretiens par mois'\n"
            response += "- 'Comparez les recruteurs sur les recrutements'"
        
        else:
            response = "Je n'ai pas compris votre demande. Voici ce que je peux vous dire:\n"
            response += f"- Total candidats contactés: {kpis['total_contacted']}\n"
            response += f"- Total entretiens: {kpis['total_interviews']}\n"
            response += f"- Total recrutements: {kpis['total_recruitments']}\n\n"
            response += "Posez une question plus précise ou demandez de l'aide pour voir ce que je peux faire."
        
        # Affichage de la réponse
        st.session_state.messages.append({"role": "assistant", "content": response})
        with st.chat_message("assistant"):
            st.markdown(response)

if __name__ == "__main__":
    main()
